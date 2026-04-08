"""
Write test execution results as an XLSX file.

Format mirrors TestCases.xlsx (Requirement ID, Test Case ID, Step #,
Step Description, Expected Output, Test Data) with an additional
Step Result column (PASS / FAIL / PARTIAL / N/A).
"""
from __future__ import annotations

from collections import defaultdict
from pathlib import Path
from typing import TYPE_CHECKING

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

if TYPE_CHECKING:
    from tools.run_logger import RunLogger

_HEADERS = [
    "Requirement ID",
    "Test Case ID",
    "Step #",
    "Step Description",
    "Expected Output",
    "Test Data",
    "Step Result",
]

_COL_WIDTHS = [15, 12, 8, 50, 40, 30, 14]

_STATUS_COLORS: dict[str, str] = {
    "PASS":    "FF90EE90",  # light green
    "FAIL":    "FFFF6B6B",  # light red
    "PARTIAL": "FFFFA500",  # orange
    "SKIPPED": "FFFFD700",  # gold
    "N/A":     "FFD3D3D3",  # light grey
}


def write_result_xlsx(logger: "RunLogger", out_path: Path) -> None:
    """Write a result XLSX for the given run into *out_path*."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Results"

    _write_header(ws)

    step_results = _compute_step_results(logger)

    if logger.test_case:
        _write_tc_rows(ws, logger.test_case, step_results)
    else:
        _write_adhoc_rows(ws, logger)

    for col_idx, width in enumerate(_COL_WIDTHS, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)


# ---------------------------------------------------------------------------
# Internal helpers
# ---------------------------------------------------------------------------

def _write_header(ws) -> None:
    for col_idx, header in enumerate(_HEADERS, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.alignment = Alignment(wrap_text=True)


def _compute_step_results(logger: "RunLogger") -> dict[int, str]:
    """Return {tcStep: aggregated_status} from the execution trace."""
    if not logger.run_result:
        return {}
    trace = logger.run_result.get("execution_trace", [])
    groups: dict[int, list[dict]] = defaultdict(list)
    for sr in trace:
        tc = sr.get("tcStep")
        if tc is not None:
            groups[int(tc)].append(sr)
    results: dict[int, str] = {}
    for step_num, entries in groups.items():
        if any(e.get("status") == "FAIL" for e in entries):
            results[step_num] = "FAIL"
        elif any(e.get("status") == "SKIPPED" for e in entries):
            results[step_num] = "PARTIAL"
        else:
            results[step_num] = "PASS"
    return results


def _write_tc_rows(ws, test_case: dict, step_results: dict[int, str]) -> None:
    req_id = test_case.get("requirementId", "")
    tc_id = test_case.get("testCaseId", "")
    for row_idx, step in enumerate(test_case.get("steps", []), 2):
        step_num = step.get("stepNumber", row_idx - 1)
        status = step_results.get(step_num, "N/A")
        row_data = [
            req_id,
            tc_id,
            step_num,
            step.get("stepDescription", ""),
            step.get("expectedOutput") or "",
            step.get("testData") or "",
            status,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 7:
                _apply_status_color(cell, status)


def _write_adhoc_rows(ws, logger: "RunLogger") -> None:
    """For ad-hoc runs without TC metadata: one row per execution trace entry."""
    if not logger.run_result:
        return
    trace = logger.run_result.get("execution_trace", [])
    for row_idx, sr in enumerate(trace, 2):
        status = sr.get("status", "N/A")
        row_data = [
            "",
            "",
            sr.get("step", row_idx - 1),
            sr.get("action", ""),
            sr.get("detail", ""),
            "",
            status,
        ]
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            if col_idx == 7:
                _apply_status_color(cell, status)


def _apply_status_color(cell, status: str) -> None:
    color = _STATUS_COLORS.get(status, _STATUS_COLORS["N/A"])
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")
