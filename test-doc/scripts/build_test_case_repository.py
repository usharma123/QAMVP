#!/usr/bin/env python3
"""Render a detailed test-case repository from a JSON spec."""

from __future__ import annotations

import argparse
import json
import sys
from dataclasses import dataclass
from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parents[2]
DEFAULT_SPEC = ROOT / "test-doc" / "test-case-repository.json"
DEFAULT_WORKBOOK_OUT = ROOT / "test_data" / "TestCases.xlsx"
DEFAULT_MARKDOWN_OUT = ROOT / "test-doc" / "09-test-case-repository.md"

HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)


@dataclass(frozen=True)
class Step:
    requirement_id: str
    step_number: int
    step_description: str
    expected_output: str
    test_data: str = ""


@dataclass(frozen=True)
class Case:
    test_case_id: str
    title: str
    objective: str
    priority: str
    suite: str
    tags: tuple[str, ...]
    steps: tuple[Step, ...]


@dataclass(frozen=True)
class RepositorySpec:
    title: str
    document_id: str
    version: str
    status: str
    source_workbook: str
    cases: tuple[Case, ...]


def load_spec(path: Path) -> RepositorySpec:
    data = json.loads(path.read_text(encoding="utf-8"))
    cases = tuple(_parse_case(raw, i) for i, raw in enumerate(data.get("cases", []), start=1))
    if not cases:
        raise ValueError(f"{path} must define at least one case")

    seen: set[str] = set()
    for case in cases:
        if case.test_case_id in seen:
            raise ValueError(f"duplicate test_case_id: {case.test_case_id}")
        seen.add(case.test_case_id)
        step_numbers = [step.step_number for step in case.steps]
        if len(step_numbers) != len(set(step_numbers)):
            raise ValueError(f"{case.test_case_id} has duplicate step_number values")

    return RepositorySpec(
        title=str(data.get("title") or "Detailed Test Case Repository"),
        document_id=str(data.get("document_id") or "TC-REPO-001"),
        version=str(data.get("version") or "0.1"),
        status=str(data.get("status") or "Generated test-case repository."),
        source_workbook=str(data.get("source_workbook") or "test_data/TestCases.xlsx"),
        cases=cases,
    )


def _parse_case(raw: dict[str, Any], ordinal: int) -> Case:
    test_case_id = _required_text(raw, "test_case_id", f"case #{ordinal}").upper()
    steps = tuple(
        _parse_step(step, test_case_id, i)
        for i, step in enumerate(raw.get("steps", []), start=1)
    )
    if not steps:
        raise ValueError(f"{test_case_id} must define at least one step")

    raw_tags = raw.get("tags", [])
    if isinstance(raw_tags, str):
        tags = tuple(t.strip() for t in raw_tags.replace(";", ",").split(",") if t.strip())
    else:
        tags = tuple(str(t).strip() for t in raw_tags if str(t).strip())

    return Case(
        test_case_id=test_case_id,
        title=_required_text(raw, "title", test_case_id),
        objective=_required_text(raw, "objective", test_case_id),
        priority=str(raw.get("priority") or "P1"),
        suite=str(raw.get("suite") or "Regression"),
        tags=tags,
        steps=tuple(sorted(steps, key=lambda step: step.step_number)),
    )


def _parse_step(raw: dict[str, Any], test_case_id: str, ordinal: int) -> Step:
    step_number = raw.get("step_number", ordinal)
    try:
        step_no = int(step_number)
    except (TypeError, ValueError) as exc:
        raise ValueError(f"{test_case_id} has invalid step_number: {step_number!r}") from exc

    return Step(
        requirement_id=_required_text(raw, "requirement_id", f"{test_case_id} step {step_no}").upper(),
        step_number=step_no,
        step_description=_required_text(raw, "step_description", f"{test_case_id} step {step_no}"),
        expected_output=str(raw.get("expected_output") or ""),
        test_data=str(raw.get("test_data") or ""),
    )


def _required_text(raw: dict[str, Any], key: str, context: str) -> str:
    value = str(raw.get(key) or "").strip()
    if not value:
        raise ValueError(f"{context} missing required field: {key}")
    return value


def style_header(ws, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: list[float]) -> None:
    for i, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = width


def build_workbook(spec: RepositorySpec, workbook_out: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "TestCases"
    headers = ["RequirementID", "TestCaseID", "StepNumber", "StepDescription", "ExpectedOutput", "TestData"]
    ws.append(headers)
    for case in spec.cases:
        for step in case.steps:
            ws.append(
                [
                    step.requirement_id,
                    case.test_case_id,
                    step.step_number,
                    step.step_description,
                    step.expected_output,
                    step.test_data,
                ]
            )
    style_header(ws, len(headers))
    set_widths(ws, [18, 14, 12, 72, 72, 64])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:F{ws.max_row}"
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    catalog = wb.create_sheet("Catalog")
    catalog_headers = ["TestCaseID", "Title", "Objective", "Priority", "Suite", "Tags", "StepCount", "Requirements"]
    catalog.append(catalog_headers)
    for case in spec.cases:
        reqs = sorted({step.requirement_id for step in case.steps})
        catalog.append(
            [
                case.test_case_id,
                case.title,
                case.objective,
                case.priority,
                case.suite,
                ", ".join(case.tags),
                len(case.steps),
                ", ".join(reqs),
            ]
        )
    style_header(catalog, len(catalog_headers))
    set_widths(catalog, [14, 56, 76, 10, 24, 36, 10, 52])
    catalog.freeze_panes = "A2"
    catalog.auto_filter.ref = f"A1:H{catalog.max_row}"
    for row in catalog.iter_rows(min_row=2):
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)

    summary = wb.create_sheet("Summary")
    summary_rows = [
        ("Metric", "Value"),
        ("Document ID", spec.document_id),
        ("Version", spec.version),
        ("Test case count", len(spec.cases)),
        ("Step count", sum(len(case.steps) for case in spec.cases)),
        ("P1 cases", sum(1 for case in spec.cases if case.priority == "P1")),
        ("Source spec", "test-doc/test-case-repository.json"),
    ]
    for row in summary_rows:
        summary.append(row)
    style_header(summary, 2)
    set_widths(summary, [28, 72])

    workbook_out.parent.mkdir(parents=True, exist_ok=True)
    wb.save(workbook_out)


def build_markdown(spec: RepositorySpec, markdown_out: Path) -> None:
    lines: list[str] = [
        f"# {spec.title}",
        "",
        f"**Document ID:** {spec.document_id}",
        f"**Version:** {spec.version}",
        f"**Status:** {spec.status}",
        "",
        "This document is generated from `test-doc/test-case-repository.json`. It gives the RAG corpus the same row-level test-case detail as the Excel inventory, while the ingestion pipeline also loads the workbook into structured SQL tables.",
        "",
        "## Inventory Summary",
        "",
        f"- Test cases: {len(spec.cases)}",
        f"- Test steps: {sum(len(case.steps) for case in spec.cases)}",
        f"- Source workbook: `{spec.source_workbook}`",
        "",
        "## Catalog",
        "",
        "| TestCaseID | Title | Priority | Suite | Requirements |",
        "|------------|-------|----------|-------|--------------|",
    ]
    for case in spec.cases:
        reqs = ", ".join(sorted({step.requirement_id for step in case.steps}))
        lines.append(
            f"| {case.test_case_id} | {md(case.title)} | {case.priority} | {case.suite} | {reqs} |"
        )

    for case in spec.cases:
        lines.extend(
            [
                "",
                f"## {case.test_case_id} - {case.title}",
                "",
                f"**Objective:** {case.objective}",
                f"**Priority:** {case.priority}",
                f"**Suite:** {case.suite}",
                f"**Tags:** {', '.join(case.tags)}",
                "",
                "| Step | RequirementID | StepDescription | ExpectedOutput | TestData |",
                "|------|---------------|-----------------|----------------|----------|",
            ]
        )
        for step in case.steps:
            lines.append(
                f"| {step.step_number} | {step.requirement_id} | {md(step.step_description)} | {md(step.expected_output)} | {md(step.test_data)} |"
            )

    markdown_out.parent.mkdir(parents=True, exist_ok=True)
    markdown_out.write_text("\n".join(lines) + "\n", encoding="utf-8")


def md(value: str) -> str:
    return value.replace("|", "\\|")


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Render TestCases.xlsx and 09-test-case-repository.md from a JSON spec."
    )
    parser.add_argument("--spec", type=Path, default=DEFAULT_SPEC)
    parser.add_argument("--workbook-out", type=Path, default=DEFAULT_WORKBOOK_OUT)
    parser.add_argument("--markdown-out", type=Path, default=DEFAULT_MARKDOWN_OUT)
    return parser.parse_args()


def main() -> None:
    args = parse_args()
    try:
        spec = load_spec(args.spec)
    except Exception as exc:
        print(f"ERROR: {exc}", file=sys.stderr)
        sys.exit(1)

    build_workbook(spec, args.workbook_out)
    build_markdown(spec, args.markdown_out)
    print(f"Wrote {args.workbook_out}")
    print(f"Wrote {args.markdown_out}")
    print(
        f"Inventory: {len(spec.cases)} test cases, "
        f"{sum(len(case.steps) for case in spec.cases)} steps"
    )


if __name__ == "__main__":
    main()
