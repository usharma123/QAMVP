"""
Write two Excel reports for a completed test run.

Report 1 — Detail Trace  (TC-001_detail_<ts>.xlsx):
    One row per sub-step in execution_trace — full drill-down for debugging.

Report 2 — TC Step Summary  (TC-001_summary_<ts>.xlsx):
    One row per TestCases.xlsx step — business-readable, mirrors TC sheet.
"""
from __future__ import annotations

import re
from collections import defaultdict
from pathlib import Path
from typing import Optional

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill

_STATUS_COLORS: dict[str, str] = {
    "PASS":    "FF90EE90",  # light green
    "FAIL":    "FFFF6B6B",  # light red
    "PARTIAL": "FFFFA500",  # orange
    "SKIPPED": "FFFFD700",  # gold
    "NOT_RUN": "FFFFFF99",  # light yellow
    "N/A":     "FFD3D3D3",  # light grey
}

_HEADER_FILL = PatternFill(start_color="FFB8CCE4", end_color="FFB8CCE4", fill_type="solid")
_HEALED_FILL = PatternFill(start_color="FFFFCC00", end_color="FFFFCC00", fill_type="solid")

# ---------------------------------------------------------------------------
# Detail Report columns
# ---------------------------------------------------------------------------
_DETAIL_HEADERS = [
    "Run ID",
    "Requirement ID",
    "Test Case ID",
    "TC Step",
    "Step Description",
    "Sub-step Index",
    "Sub-step Label",
    "Parent Action",
    "Action",
    "Status",
    "Expected Output",
    "Actual Output",
    "Error Detail",
    "Self-Healed",
]
_DETAIL_WIDTHS = [22, 14, 12, 10, 45, 14, 22, 20, 20, 12, 35, 35, 60, 12]

# ---------------------------------------------------------------------------
# Summary Report columns
# ---------------------------------------------------------------------------
_SUMMARY_HEADERS = [
    "Requirement ID",
    "Test Case ID",
    "Step #",
    "Step Description",
    "Expected Output",
    "Test Data",
    "Actual Output",
    "Step Result",
    "Failure Detail",
    "Self-Healed",
    "Sub-steps",
]
_SUMMARY_WIDTHS = [15, 12, 8, 50, 40, 30, 40, 14, 60, 12, 22]

_ASSERT_ACTIONS = {"ASSERT_TEXT", "ASSERT_CONTAINS", "ASSERT_VARIABLE", "ASSERT_VISIBLE"}
_READ_ACTIONS = {"READ_TEXT", "READ_ATTRIBUTE"}


# ---------------------------------------------------------------------------
# Public API
# ---------------------------------------------------------------------------

def write_both_reports(
    result: dict,
    test_case: Optional[dict],
    script: Optional[dict],
    out_dir: Path,
    timestamp: str,
) -> tuple[Path, Path]:
    """Write detail + summary XLSX reports. Returns (detail_path, summary_path)."""
    tc_id = result.get("testCaseId") or (test_case or {}).get("testCaseId", "unknown")
    out_dir.mkdir(parents=True, exist_ok=True)
    detail_path = out_dir / f"{tc_id}_detail_{timestamp}.xlsx"
    summary_path = out_dir / f"{tc_id}_summary_{timestamp}.xlsx"
    _write_detail_report(result, test_case, script, detail_path)
    _write_summary_report(result, test_case, script, summary_path)
    return detail_path, summary_path


# ---------------------------------------------------------------------------
# Report 1 — Detail Trace
# ---------------------------------------------------------------------------

def _write_detail_report(
    result: dict,
    test_case: Optional[dict],
    script: Optional[dict],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Execution Trace"
    _write_header_row(ws, _DETAIL_HEADERS, _DETAIL_WIDTHS)

    trace = result.get("execution_trace", [])
    healed_steps = _build_healed_steps(result.get("help_exchanges", []))
    tc_step_meta = _build_tc_step_meta(test_case)
    script_step_map = _build_script_step_map(script)

    req_id = result.get("requirementId", "")
    tc_id = result.get("testCaseId", "")
    run_id = f"{tc_id}_{result.get('_run_timestamp', '')}" if tc_id else "unknown"

    for row_idx, sr in enumerate(trace, 2):
        action = sr.get("action", "")
        status = sr.get("status", "N/A")
        tc_step = sr.get("tcStep")
        step_idx = sr.get("step", "")
        parent = sr.get("parentAction", "")

        meta = tc_step_meta.get(int(tc_step)) if tc_step is not None else {}
        desc = (meta or {}).get("stepDescription", "")
        expected = (meta or {}).get("expectedOutput") or ""

        # Only look up script step for primitive (non-macro) steps
        script_step = None
        if not parent and tc_step is not None:
            script_step = script_step_map.get((int(tc_step), action.upper()))

        actual = _extract_actual_output(action, status, sr.get("detail", ""), script_step)

        row = [
            run_id,
            req_id,
            tc_id,
            tc_step if tc_step is not None else "",
            desc,
            step_idx,
            sr.get("subStepLabel", ""),
            parent,
            action,
            status,
            expected,
            actual,
            sr.get("detail", ""),
            "YES" if step_idx in healed_steps else "NO",
        ]
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)
            if col_idx == 10:  # Status
                _apply_status_color(cell, status)
            if col_idx == 14 and value == "YES":  # Self-Healed
                cell.fill = _HEALED_FILL

    # Self-Healing Log sheet (only when exchanges exist)
    exchanges = result.get("help_exchanges", [])
    if exchanges:
        _write_healing_sheet(wb, exchanges, trace)

    wb.save(out_path)


def _write_healing_sheet(wb, exchanges: list[dict], trace: list[dict]) -> None:
    import json as _json
    ws = wb.create_sheet("Self-Healing Log")
    headers = [
        "Exchange #", "Step", "Action", "Locator",
        "Error", "Current URL", "Solution Action", "Solution Params", "Outcome",
    ]
    widths = [12, 8, 18, 25, 70, 30, 16, 30, 10]
    _write_header_row(ws, headers, widths)

    trace_by_step = {e.get("step"): e for e in trace}

    for i in range(0, len(exchanges), 2):
        req = exchanges[i] if i < len(exchanges) else {}
        sol = exchanges[i + 1] if i + 1 < len(exchanges) else {}
        step_idx = req.get("step")
        outcome = trace_by_step.get(step_idx, {}).get("status", "?")
        row = [
            i // 2 + 1,
            step_idx,
            req.get("action", ""),
            req.get("locator", ""),
            req.get("error", ""),
            req.get("currentUrl", ""),
            sol.get("action", ""),
            _json.dumps(sol.get("params", {})),
            outcome,
        ]
        for col_idx, value in enumerate(row, 1):
            cell = ws.cell(row=i // 2 + 2, column=col_idx, value=value)
            cell.alignment = Alignment(wrap_text=True)


# ---------------------------------------------------------------------------
# Report 2 — TC Step Summary
# ---------------------------------------------------------------------------

def _write_summary_report(
    result: dict,
    test_case: Optional[dict],
    script: Optional[dict],
    out_path: Path,
) -> None:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "TC Step Summary"
    _write_header_row(ws, _SUMMARY_HEADERS, _SUMMARY_WIDTHS)

    trace = result.get("execution_trace", [])
    healed_steps = _build_healed_steps(result.get("help_exchanges", []))
    script_step_map = _build_script_step_map(script)

    groups: dict[int, list[dict]] = defaultdict(list)
    for sr in trace:
        tc = sr.get("tcStep")
        if tc is not None:
            groups[int(tc)].append(sr)

    req_id = result.get("requirementId", "")
    tc_id = result.get("testCaseId", "")
    steps = (test_case or {}).get("steps", [])

    if steps:
        for row_idx, step in enumerate(steps, 2):
            step_num = int(step.get("stepNumber", row_idx - 1))
            sub_steps = groups.get(step_num, [])

            if not sub_steps:
                status = "NOT_RUN"
                actual = ""
                failure_detail = _not_run_note(step_num, groups)
                healed = "NO"
                sub_summary = ""
            else:
                status = _aggregate_status(sub_steps)
                actual = _extract_actual_for_tc_step(sub_steps, script_step_map)
                failure_detail = _first_failure_detail(sub_steps)
                healed = "YES" if any(e.get("step") in healed_steps for e in sub_steps) else "NO"
                sub_summary = _sub_step_summary(sub_steps)

            row = [
                req_id,
                tc_id,
                step_num,
                step.get("stepDescription", ""),
                step.get("expectedOutput") or "",
                step.get("testData") or "",
                actual,
                status,
                failure_detail,
                healed,
                sub_summary,
            ]
            _write_summary_row(ws, row_idx, row, status_col=8, healed_col=10)
    else:
        # No TC metadata — one row per tcStep group
        for row_idx, (step_num, sub_steps) in enumerate(sorted(groups.items()), 2):
            status = _aggregate_status(sub_steps)
            row = [
                req_id, tc_id, step_num, "", "", "",
                _extract_actual_for_tc_step(sub_steps, script_step_map),
                status,
                _first_failure_detail(sub_steps),
                "YES" if any(e.get("step") in healed_steps for e in sub_steps) else "NO",
                _sub_step_summary(sub_steps),
            ]
            _write_summary_row(ws, row_idx, row, status_col=8, healed_col=10)

    wb.save(out_path)


def _write_summary_row(ws, row_idx: int, row: list, status_col: int, healed_col: int) -> None:
    for col_idx, value in enumerate(row, 1):
        cell = ws.cell(row=row_idx, column=col_idx, value=value)
        cell.alignment = Alignment(wrap_text=True)
        if col_idx == status_col:
            _apply_status_color(cell, str(value))
        if col_idx == healed_col and value == "YES":
            cell.fill = _HEALED_FILL


# ---------------------------------------------------------------------------
# Actual output extraction
# ---------------------------------------------------------------------------

def _extract_actual_output(
    action: str,
    status: str,
    detail: str,
    script_step: Optional[dict],
) -> str:
    """Derive the best-available actual output for a single sub-step."""
    action_upper = action.upper()

    if action_upper in _ASSERT_ACTIONS:
        if status == "PASS":
            return "= Expected (PASS)"
        if detail:
            if action_upper == "ASSERT_VISIBLE":
                return "not visible"
            m = re.search(r"got ['\"](.+?)['\"]", detail)
            if m:
                return m.group(1)
            return detail[:120]

    if action_upper in _READ_ACTIONS:
        var_name = (script_step or {}).get("output", "")
        label = f"→ {var_name}" if var_name else ""
        return f"Captured {label} (value not recorded)"

    return ""


def _extract_actual_for_tc_step(
    sub_steps: list[dict],
    script_step_map: dict,
) -> str:
    """Best actual output for a whole TC step — prefer failing assertions, then passing, then reads."""
    for sr in sub_steps:
        if sr.get("action", "").upper() in _ASSERT_ACTIONS and sr.get("status") == "FAIL":
            return _extract_actual_output(
                sr["action"], "FAIL", sr.get("detail", ""), None
            )
    for sr in sub_steps:
        if sr.get("action", "").upper() in _ASSERT_ACTIONS and sr.get("status") == "PASS":
            return "= Expected (PASS)"
    for sr in sub_steps:
        if sr.get("action", "").upper() in _READ_ACTIONS:
            tc = sr.get("tcStep")
            script_step = None
            if tc is not None and not sr.get("parentAction"):
                script_step = script_step_map.get((int(tc), sr["action"].upper()))
            return _extract_actual_output(
                sr["action"], sr.get("status", ""), sr.get("detail", ""), script_step
            )
    return ""


# ---------------------------------------------------------------------------
# Status aggregation
# ---------------------------------------------------------------------------

def _aggregate_status(sub_steps: list[dict]) -> str:
    statuses = {e.get("status", "") for e in sub_steps}
    if "FAIL" in statuses:
        return "FAIL"
    if "SKIPPED" in statuses and "PASS" in statuses:
        return "PARTIAL"
    if statuses <= {"SKIPPED"}:
        return "SKIPPED"
    if "PASS" in statuses:
        return "PASS"
    return "N/A"


def _first_failure_detail(sub_steps: list[dict]) -> str:
    for sr in sub_steps:
        if sr.get("status") == "FAIL" and sr.get("detail"):
            return sr["detail"][:200]
    return ""


def _sub_step_summary(sub_steps: list[dict]) -> str:
    passed  = sum(1 for e in sub_steps if e.get("status") == "PASS")
    failed  = sum(1 for e in sub_steps if e.get("status") == "FAIL")
    skipped = sum(1 for e in sub_steps if e.get("status") == "SKIPPED")
    total   = len(sub_steps)
    parts = []
    if passed:  parts.append(f"{passed} passed")
    if failed:  parts.append(f"{failed} failed")
    if skipped: parts.append(f"{skipped} skipped")
    return f"{', '.join(parts)} / {total}"


def _not_run_note(step_num: int, groups: dict[int, list]) -> str:
    next_steps = sorted(s for s in groups if s > step_num)
    if next_steps:
        return f"Not mapped in script — tcStep {next_steps[0]} may cover this intent"
    return "Not mapped in script"


# ---------------------------------------------------------------------------
# Shared helpers
# ---------------------------------------------------------------------------

def _write_header_row(ws, headers: list[str], widths: list[int]) -> None:
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)
        cell.fill = _HEADER_FILL
        cell.alignment = Alignment(wrap_text=True, horizontal="center")
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = width


def _apply_status_color(cell, status: str) -> None:
    color = _STATUS_COLORS.get(status, _STATUS_COLORS["N/A"])
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type="solid")


def _build_healed_steps(exchanges: list[dict]) -> set:
    return {
        exchanges[i].get("step")
        for i in range(0, len(exchanges), 2)
        if i < len(exchanges) and exchanges[i].get("step") is not None
    }


def _build_tc_step_meta(test_case: Optional[dict]) -> dict[int, dict]:
    if not test_case:
        return {}
    return {
        int(s["stepNumber"]): s
        for s in test_case.get("steps", [])
        if s.get("stepNumber") is not None
    }


def _build_script_step_map(script: Optional[dict]) -> dict[tuple, dict]:
    """Map (tcStep, ACTION) -> script step for primitive steps."""
    if not script:
        return {}
    result: dict[tuple, dict] = {}
    for step in script.get("steps", []):
        tc = step.get("tcStep")
        action = step.get("action", "").upper()
        if tc is not None and action:
            result[(int(tc), action)] = step
    return result
