"""
Test data defaults — reads testdatadefaults.xlsx for configurable values.

When the test step does not explicitly provide values (ticker, credentials,
quantity, etc.), the orchestrator infers them from this sheet.
"""

from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
TEST_DATA_DEFAULTS_XLSX = REPO_ROOT / "test_data" / "testdatadefaults.xlsx"

_FALLBACK_DEFAULTS = {
    "maker_username": "admin",
    "maker_password": "admin",
    "checker_username": "checker",
    "checker_password": "chscker@123",
    "sector_options": "Technology,Financials",
    "ticker_technology": "AAPL,MSFT,NVDA",
    "ticker_financials": "JPM,BAC,V",
    "quantity_min": "1",
    "quantity_max": "500",
    "account_type_options": "Cash,Margin",
    "time_in_force_default": "Day Order",
}


def load_test_data_defaults() -> dict[str, str]:
    """Read defaults from testdatadefaults.xlsx. Returns {param: value}.
    Falls back to hardcoded values if file does not exist."""
    if not TEST_DATA_DEFAULTS_XLSX.exists():
        return dict(_FALLBACK_DEFAULTS)

    wb = openpyxl.load_workbook(TEST_DATA_DEFAULTS_XLSX, read_only=True, data_only=True)
    result: dict[str, str] = dict(_FALLBACK_DEFAULTS)

    sheet = wb["defaults"] if "defaults" in wb.sheetnames else wb.active
    for row in sheet.iter_rows(min_row=2, values_only=True):
        param = str(row[0] or "").strip().lower().replace(" ", "_")
        value = str(row[1] or "").strip()
        if param:
            result[param] = value

    wb.close()
    return result


def format_test_data_defaults_for_prompt(defaults: dict[str, str]) -> str:
    """Convert defaults dict into markdown bullet list for the step prompt."""
    maker_user = defaults.get("maker_username", "admin")
    maker_pass = defaults.get("maker_password", "admin")
    checker_user = defaults.get("checker_username", "checker")
    checker_pass = defaults.get("checker_password", "chscker@123")
    sector_opts = defaults.get("sector_options", "Technology,Financials")
    ticker_tech = defaults.get("ticker_technology", "AAPL,MSFT,NVDA")
    ticker_fin = defaults.get("ticker_financials", "JPM,BAC,V")
    qty_min = defaults.get("quantity_min", "1")
    qty_max = defaults.get("quantity_max", "500")
    acct_opts = defaults.get("account_type_options", "Cash,Margin")
    tif = defaults.get("time_in_force_default", "Day Order")

    lines = [
        f"- **Maker credentials**: `{maker_user}` / `{maker_pass}`.",
        f"- **Checker credentials**: `{checker_user}` / `{checker_pass}`.",
        f"- **Sector**: Randomly pick from `{sector_opts}`.",
        f"- **Ticker**: Pick a valid ticker for the sector (Technology: {ticker_tech}; Financials: {ticker_fin}).",
        f"- **Quantity**: Generate a realistic integer between {qty_min} and {qty_max}.",
        f"- **Account Type**: Pick from `{acct_opts}`.",
        f"- **Time in Force**: Use `{tif}` unless the tester specifies GTC.",
    ]
    return "\n".join(lines)
