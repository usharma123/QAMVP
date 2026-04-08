"""
Shared locator store — reads locators.xlsx (one sheet per page).

Every Python module that needs locator data imports from here instead of
reading the file directly.  This keeps the xlsx parsing logic in one place.
"""

from pathlib import Path

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
LOCATORS_XLSX = REPO_ROOT / "test_data" / "locators.xlsx"


def load_all_locators() -> list[dict[str, str]]:
    """Return a flat list of {element_name, xpath, page} dicts from all sheets."""
    if not LOCATORS_XLSX.exists():
        return []
    wb = openpyxl.load_workbook(LOCATORS_XLSX, read_only=True, data_only=True)
    rows: list[dict[str, str]] = []
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        headers = [str(c.value or "").strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))]
        name_idx = headers.index("element_name") if "element_name" in headers else 0
        xpath_idx = headers.index("xpath") if "xpath" in headers else 1
        for row in ws.iter_rows(min_row=2, values_only=True):
            ename = str(row[name_idx] or "").strip()
            xpath = str(row[xpath_idx] or "").strip()
            if ename:
                rows.append({
                    "element_name": ename,
                    "xpath": xpath,
                    "page": sheet_name,
                })
    wb.close()
    return rows


def load_locator_page_map() -> dict[str, str]:
    """Return {element_name: page} for all locators."""
    return {r["element_name"]: r["page"] for r in load_all_locators() if r.get("element_name")}


def load_page_elements() -> dict[str, set[str]]:
    """Return {page: set of element names}."""
    pages: dict[str, set[str]] = {}
    for r in load_all_locators():
        page = r.get("page", "")
        if page:
            pages.setdefault(page, set()).add(r["element_name"])
    return pages


def load_locators_for_page(page: str) -> list[dict[str, str]]:
    """Return locators for a single page/sheet."""
    return [r for r in load_all_locators() if r.get("page") == page]


def load_locators_for_pages(pages: list[str]) -> list[dict[str, str]]:
    """Return the union of locators for multiple pages/sheets."""
    page_set = set(pages)
    return [r for r in load_all_locators() if r.get("page") in page_set]


def update_locator(element_name: str, new_xpath: str) -> None:
    """Update a single locator's xpath in-place across all sheets."""
    if not LOCATORS_XLSX.exists():
        return
    wb = openpyxl.load_workbook(LOCATORS_XLSX)
    for ws in wb.worksheets:
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        name_idx = headers.index("element_name") if "element_name" in headers else 0
        xpath_idx = headers.index("xpath") if "xpath" in headers else 1
        for row in ws.iter_rows(min_row=2):
            if str(row[name_idx].value or "").strip() == element_name:
                row[xpath_idx].value = new_xpath
                wb.save(LOCATORS_XLSX)
                wb.close()
                return
    wb.close()


def rename_locator(old_name: str, new_name: str, new_xpath: str) -> bool:
    """Rename a locator (element_name + xpath) in-place. Returns True if found."""
    if not LOCATORS_XLSX.exists():
        return False
    wb = openpyxl.load_workbook(LOCATORS_XLSX)
    for ws in wb.worksheets:
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        name_idx = headers.index("element_name") if "element_name" in headers else 0
        xpath_idx = headers.index("xpath") if "xpath" in headers else 1
        for row in ws.iter_rows(min_row=2):
            if str(row[name_idx].value or "").strip() == old_name:
                row[name_idx].value = new_name
                row[xpath_idx].value = new_xpath
                wb.save(LOCATORS_XLSX)
                wb.close()
                return True
    wb.close()
    return False


def add_locator(element_name: str, xpath: str, page: str) -> None:
    """Append a locator row to the sheet named `page` (creates the sheet if missing)."""
    if not LOCATORS_XLSX.exists():
        return
    wb = openpyxl.load_workbook(LOCATORS_XLSX)
    if page in wb.sheetnames:
        ws = wb[page]
    else:
        ws = wb.create_sheet(page)
        ws.append(["element_name", "xpath"])
    ws.append([element_name, xpath])
    wb.save(LOCATORS_XLSX)
    wb.close()


def deduplicate_locators() -> int:
    """Remove duplicate element_name rows across all sheets. Keeps first occurrence."""
    if not LOCATORS_XLSX.exists():
        return 0
    wb = openpyxl.load_workbook(LOCATORS_XLSX)
    seen: set[str] = set()
    removed = 0
    for ws in wb.worksheets:
        headers = [str(c.value or "").strip().lower() for c in ws[1]]
        name_idx = headers.index("element_name") if "element_name" in headers else 0
        rows_to_delete: list[int] = []
        for row in ws.iter_rows(min_row=2):
            ename = str(row[name_idx].value or "").strip()
            if not ename:
                continue
            if ename in seen:
                rows_to_delete.append(row[0].row)
            else:
                seen.add(ename)
        for row_num in reversed(rows_to_delete):
            ws.delete_rows(row_num)
            removed += 1
    if removed:
        wb.save(LOCATORS_XLSX)
    wb.close()
    return removed
