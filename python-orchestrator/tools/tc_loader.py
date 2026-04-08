"""
Load structured test cases from TestCases.xlsx for traceability and LLM input.
"""

from __future__ import annotations

import re
from pathlib import Path
from typing import Any, Optional

import openpyxl

REPO_ROOT = Path(__file__).resolve().parent.parent.parent
TEST_CASES_XLSX = REPO_ROOT / "test_data" / "TestCases.xlsx"


def _normalize_tc_id(raw: str) -> str:
    """Accept TC-001, tc-001, TC001 -> TC-001 style when possible."""
    raw = raw.strip().upper()
    if re.match(r"^TC-\d+$", raw):
        return raw
    m = re.match(r"^TC(\d+)$", raw)
    if m:
        return f"TC-{int(m.group(1)):03d}"
    return raw.strip()


def list_test_cases() -> list[dict[str, str]]:
    if not TEST_CASES_XLSX.exists():
        raise FileNotFoundError(f"TestCases.xlsx not found: {TEST_CASES_XLSX}")
    wb = openpyxl.load_workbook(TEST_CASES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    seen: dict[str, str] = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        req, tc = row[0], row[1]
        if not tc:
            continue
        tid = str(tc).strip()
        seen[tid] = str(req).strip() if req else ""
    return [{"testCaseId": k, "requirementId": v} for k, v in sorted(seen.items())]


def load_test_case(tc_id: str) -> Optional[dict[str, Any]]:
    """Return a test case dict or None if not found."""
    if not TEST_CASES_XLSX.exists():
        raise FileNotFoundError(f"TestCases.xlsx not found: {TEST_CASES_XLSX}")
    want = _normalize_tc_id(tc_id)
    wb = openpyxl.load_workbook(TEST_CASES_XLSX, read_only=True, data_only=True)
    ws = wb.active
    collected: list[dict[str, Any]] = []
    req_id = ""
    sheet_tc_id = ""

    for row in ws.iter_rows(min_row=2, values_only=True):
        req, tc = row[0], row[1]
        if not tc:
            continue
        stc = str(tc).strip()
        if _normalize_tc_id(stc) != want:
            continue
        if not collected:
            req_id = str(req).strip() if req else ""
            sheet_tc_id = stc
        step_num, desc, exp_out, test_data = row[2], row[3], row[4], row[5]
        collected.append(
            {
                "stepNumber": int(step_num) if step_num is not None else 0,
                "stepDescription": str(desc).strip() if desc else "",
                "expectedOutput": exp_out,
                "testData": test_data,
            }
        )

    if not collected:
        return None

    collected.sort(key=lambda s: s["stepNumber"])
    return {
        "requirementId": req_id,
        "testCaseId": sheet_tc_id,
        "steps": collected,
    }


def format_for_llm(tc: dict[str, Any]) -> str:
    """Format spreadsheet test case as numbered NL for the LLM."""
    lines: list[str] = []
    for s in tc.get("steps", []):
        n = s.get("stepNumber", 0)
        desc = s.get("stepDescription", "")
        line = f"{n}. {desc}"
        td = s.get("testData")
        if td:
            line += f"  [TestData: {td}]"
        eo = s.get("expectedOutput")
        if eo:
            line += f"  [ExpectedOutput: {eo}]"
        lines.append(line)
    return "\n".join(lines)
