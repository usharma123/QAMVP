"""Load detailed test-case rows from TestCases.xlsx."""

from __future__ import annotations

from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path

from common import RawChunk

REQUIRED_HEADERS = [
    "RequirementID",
    "TestCaseID",
    "StepNumber",
    "StepDescription",
    "ExpectedOutput",
    "TestData",
]


@dataclass(frozen=True)
class TestCaseStep:
    requirement_id: str
    test_case_id: str
    step_number: int
    step_description: str
    expected_output: str
    test_data: str
    source_row: int


@dataclass(frozen=True)
class TestCaseRecord:
    test_case_id: str
    requirement_ids: list[str]
    title: str
    objective: str
    priority: str
    suite: str
    tags: list[str]
    steps: list[TestCaseStep]


def _cell_text(value: object) -> str:
    return "" if value is None else str(value).strip()


def load_test_case_records(path: Path) -> list[TestCaseRecord]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise RuntimeError("openpyxl is required to ingest TestCases.xlsx") from exc

    wb = load_workbook(path, data_only=True)
    if "TestCases" not in wb.sheetnames:
        raise RuntimeError(f"{path} must contain a 'TestCases' sheet")
    ws = wb["TestCases"]

    headers = [_cell_text(c.value) for c in ws[1]]
    missing = [h for h in REQUIRED_HEADERS if h not in headers]
    if missing:
        raise RuntimeError(f"{path} is missing required headers: {', '.join(missing)}")

    idx = {h: headers.index(h) for h in headers}
    rows_by_case: dict[str, list[TestCaseStep]] = defaultdict(list)
    meta_by_case: dict[str, dict[str, str]] = _load_catalog_metadata(wb)

    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [_cell_text(v) for v in row]
        if not any(values):
            continue

        test_case_id = values[idx["TestCaseID"]].upper()
        if not test_case_id:
            continue

        raw_step = values[idx["StepNumber"]]
        try:
            step_number = int(float(raw_step))
        except ValueError as exc:
            raise RuntimeError(
                f"{path} row {row_number} has invalid StepNumber {raw_step!r}"
            ) from exc

        step = TestCaseStep(
            requirement_id=values[idx["RequirementID"]].upper(),
            test_case_id=test_case_id,
            step_number=step_number,
            step_description=values[idx["StepDescription"]],
            expected_output=values[idx["ExpectedOutput"]],
            test_data=values[idx["TestData"]],
            source_row=row_number,
        )
        rows_by_case[test_case_id].append(step)

        for optional in ("Title", "Objective", "Priority", "Suite", "Tags"):
            if optional in idx and values[idx[optional]]:
                meta_by_case[test_case_id][optional] = values[idx[optional]]

    records: list[TestCaseRecord] = []
    for test_case_id in sorted(rows_by_case):
        steps = sorted(rows_by_case[test_case_id], key=lambda s: s.step_number)
        requirement_ids = sorted({s.requirement_id for s in steps if s.requirement_id})
        meta = meta_by_case[test_case_id]
        title = meta.get("Title") or _title_from_steps(test_case_id, steps)
        objective = meta.get("Objective") or _objective_from_steps(steps)
        priority = meta.get("Priority") or "P1"
        suite = meta.get("Suite") or _suite_from_id(test_case_id)
        tags = [
            t.strip()
            for t in meta.get("Tags", "").replace(";", ",").split(",")
            if t.strip()
        ]
        records.append(
            TestCaseRecord(
                test_case_id=test_case_id,
                requirement_ids=requirement_ids,
                title=title,
                objective=objective,
                priority=priority,
                suite=suite,
                tags=tags,
                steps=steps,
            )
        )
    return records


def _load_catalog_metadata(wb: object) -> dict[str, dict[str, str]]:
    if "Catalog" not in wb.sheetnames:
        return defaultdict(dict)
    ws = wb["Catalog"]
    headers = [_cell_text(c.value) for c in ws[1]]
    if "TestCaseID" not in headers:
        return defaultdict(dict)
    idx = {h: headers.index(h) for h in headers}
    out: dict[str, dict[str, str]] = defaultdict(dict)
    for row in ws.iter_rows(min_row=2, values_only=True):
        values = [_cell_text(v) for v in row]
        if not any(values):
            continue
        test_case_id = values[idx["TestCaseID"]].upper()
        if not test_case_id:
            continue
        for optional in ("Title", "Objective", "Priority", "Suite", "Tags"):
            if optional in idx and values[idx[optional]]:
                out[test_case_id][optional] = values[idx[optional]]
    return out


def records_to_chunks(records: list[TestCaseRecord]) -> list[RawChunk]:
    chunks: list[RawChunk] = []
    for record in records:
        lines = [
            f"TestCaseID: {record.test_case_id}",
            f"Title: {record.title}",
            f"Objective: {record.objective}",
            f"RequirementIDs: {', '.join(record.requirement_ids)}",
            f"Priority: {record.priority}",
            f"Suite: {record.suite}",
            f"Tags: {', '.join(record.tags)}",
            "",
            "| Step | RequirementID | StepDescription | ExpectedOutput | TestData |",
            "|------|---------------|-----------------|----------------|----------|",
        ]
        for step in record.steps:
            lines.append(
                "| "
                + " | ".join(
                    [
                        str(step.step_number),
                        step.requirement_id,
                        _md_cell(step.step_description),
                        _md_cell(step.expected_output),
                        _md_cell(step.test_data),
                    ]
                )
                + " |"
            )
        chunks.append(
            RawChunk(
                heading_path=f"Test Case Repository > {record.test_case_id}",
                content="\n".join(lines),
                metadata={
                    "test_case_id": record.test_case_id,
                    "requirement_ids": record.requirement_ids,
                    "priority": record.priority,
                    "suite": record.suite,
                    "tags": record.tags,
                    "step_count": len(record.steps),
                },
                urls=[],
            )
        )
    return chunks


def _md_cell(value: str) -> str:
    return value.replace("|", "\\|").replace("\n", " ")


def _title_from_steps(test_case_id: str, steps: list[TestCaseStep]) -> str:
    for step in steps:
        if step.step_description:
            return f"{test_case_id} - {step.step_description[:80]}"
    return test_case_id


def _objective_from_steps(steps: list[TestCaseStep]) -> str:
    reqs = sorted({s.requirement_id for s in steps if s.requirement_id})
    return f"Verify {'/'.join(reqs)} across {len(steps)} executable steps."


def _suite_from_id(test_case_id: str) -> str:
    if test_case_id in {"TC-001", "TC-002"}:
        return "S-1 Smoke"
    if test_case_id in {"TC-007", "TC-008"}:
        return "S-3 Negative"
    return "S-4 Detailed Regression"
