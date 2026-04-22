#!/usr/bin/env python3
"""Generate test-doc/SDLC-Test-Tracker.xlsx — SDLC gates, requirements, tickets, execution."""
from __future__ import annotations

import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter
except ImportError:
    print("Install openpyxl: pip install openpyxl", file=sys.stderr)
    sys.exit(1)

ROOT = Path(__file__).resolve().parents[1]
OUT = ROOT / "SDLC-Test-Tracker.xlsx"

HEADER_FILL = PatternFill("solid", fgColor="FF4472C4")
HEADER_FONT = Font(color="FFFFFFFF", bold=True)


def style_header(ws, max_col: int) -> None:
    for c in range(1, max_col + 1):
        cell = ws.cell(row=1, column=c)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def set_widths(ws, widths: list[float]) -> None:
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def sheet_instructions(wb: openpyxl.Workbook) -> None:
    ws = wb.active
    ws.title = "Instructions"
    ws["A1"] = "SDLC-Test-Tracker — QAMVP Mock Trading SUT"
    ws["A2"] = "Source Markdown: test-doc/README.md and 01–08 documents."
    ws["A3"] = "REQ IDs match 02-functional-requirements-specification.md; BR IDs match 01-business-requirements-document.md."
    ws["A4"] = "Regenerate: python test-doc/scripts/build_sdlc_tracker.py"
    ws["A5"] = "Word exports: ./test-doc/scripts/export-docx.sh"


def sheet_gates(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Gates")
    headers = [
        "Phase",
        "Gate_ID",
        "Description",
        "Owner",
        "Planned_Date",
        "Actual_Date",
        "Status",
        "Notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    rows = [
        ("Initiate", "G-INIT", "Project charter / PoC kickoff", "Sponsor", "", "", "Planned", ""),
        ("Requirements", "G-REQ", "BRD+FRS baselined", "PO", "", "", "Planned", ""),
        ("Design", "G-DES", "TDS+RTM approved", "QA Lead", "", "", "Planned", ""),
        ("Build", "G-BLD", "SUT build on localhost", "Dev", "", "", "Planned", ""),
        ("TestDesign", "G-TD", "Test plan active", "QA Lead", "", "", "Planned", ""),
        ("TestExec", "G-TE", "Smoke+regression", "QA", "", "", "Planned", ""),
        ("UAT", "G-UAT", "Sponsor demo", "Sponsor", "", "", "Planned", ""),
        ("Release", "G-REL", "PoC closure", "Sponsor", "", "", "Planned", ""),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    set_widths(ws, [14, 10, 40, 12, 14, 14, 12, 30])
    ws.freeze_panes = "A2"


def sheet_requirements(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Requirements")
    headers = [
        "REQ_ID",
        "REQ_Type",
        "Title",
        "BR_ID",
        "BRD_Section",
        "FRS_Section",
        "Priority",
        "Dev_Status",
        "Test_Status",
        "Last_Notes",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    data = [
        ("REQ-FR-000", "FR", "SPA localhost shell", "BR-010", "§4", "§2", "P2", "Done", "Planned", ""),
        ("REQ-FR-001", "FR", "Login fields present", "BR-001", "§6", "§3.1", "P1", "Done", "Planned", ""),
        ("REQ-FR-002", "FR", "Maker login", "BR-001", "§6", "§3", "P1", "Done", "Planned", ""),
        ("REQ-FR-003", "FR", "Checker login", "BR-001", "§6", "§3", "P1", "Done", "Planned", ""),
        ("REQ-FR-004", "FR", "Invalid login error", "BR-001", "§7", "§3", "P1", "Done", "Planned", ""),
        ("REQ-FR-005", "FR", "Route guard unauthenticated", "BR-001", "§7", "§3", "P1", "Done", "Planned", ""),
        ("REQ-FR-010", "FR", "Navbar persistent", "BR-011", "§6", "§4.1", "P1", "Done", "Planned", ""),
        ("REQ-FR-011", "FR", "User role display", "BR-006", "§6", "§4.1", "P1", "Done", "Planned", ""),
        ("REQ-FR-012", "FR", "Trading click menu", "BR-011", "§6", "§4.1", "P1", "Done", "Planned", ""),
        ("REQ-FR-013", "FR", "Admin menu users", "BR-013", "§6", "§4.1", "P1", "Done", "Planned", ""),
        ("REQ-FR-014", "FR", "Logout", "BR-012", "§7", "§4.1", "P1", "Done", "Planned", ""),
        ("REQ-FR-020", "FR", "Dashboard approved only", "BR-005", "§7", "§5", "P1", "Done", "Planned", ""),
        ("REQ-FR-021", "FR", "Dashboard columns", "BR-005", "§7", "§5", "P1", "Done", "Planned", ""),
        ("REQ-FR-022", "FR", "Total trades count", "BR-005", "§7", "§5", "P1", "Done", "Planned", ""),
        ("REQ-FR-023", "FR", "Dashboard loading", "BR-008", "§9", "§5", "P1", "Done", "Planned", ""),
        ("REQ-FR-030", "FR", "New trade fields", "BR-002", "§7", "§6", "P1", "Done", "Planned", ""),
        ("REQ-FR-031", "FR", "Async price total", "BR-002", "§7", "§6", "P1", "Done", "Planned", ""),
        ("REQ-FR-032", "FR", "GTC expiration", "BR-002", "§7", "§6", "P1", "Done", "Planned", ""),
        ("REQ-FR-033", "FR", "Submit trade", "BR-002", "§7", "§6", "P1", "Done", "Planned", ""),
        ("REQ-FR-034", "FR", "Submit toast", "BR-002", "§7", "§6", "P1", "Done", "Planned", ""),
        ("REQ-FR-035", "FR", "Validation", "BR-002", "§6", "§6", "P2", "Done", "Planned", ""),
        ("REQ-FR-040", "FR", "Queue pending filter", "BR-003", "§7", "§7", "P1", "Done", "Planned", ""),
        ("REQ-FR-041", "FR", "Queue columns", "BR-004", "§7", "§7", "P1", "Done", "Planned", ""),
        ("REQ-FR-042", "FR", "Pending count", "BR-004", "§7", "§7", "P1", "Done", "Planned", ""),
        ("REQ-FR-043", "FR", "Per-row approve", "BR-004", "§7", "§7", "P1", "Done", "Planned", ""),
        ("REQ-FR-044", "FR", "Approve toast", "BR-004", "§7", "§7", "P1", "Done", "Planned", ""),
        ("REQ-FR-045", "FR", "Post-approve state", "BR-003", "§7", "§7", "P1", "Done", "Planned", ""),
        ("REQ-FR-050", "FR", "Trade list columns", "BR-005", "§6", "§8", "P2", "Done", "Planned", ""),
        ("REQ-FR-051", "FR", "Trade list summary bar", "BR-005", "§6", "§8", "P2", "Done", "Planned", ""),
        ("REQ-FR-060", "FR", "Admin placeholder", "BR-013", "§6", "§9", "P3", "Done", "Planned", ""),
        ("REQ-NFR-001", "NFR", "Local operability", "BR-010", "§4", "§10", "P1", "Done", "Planned", ""),
        ("REQ-NFR-002", "NFR", "data-testid hooks", "BR-009", "§6", "§10", "P1", "Done", "Planned", ""),
        ("REQ-NFR-003", "NFR", "Responsive qualitative", "BR-010", "§4", "§10", "P3", "Done", "Planned", ""),
        ("REQ-NFR-004", "NFR", "Error UX", "BR-008", "§9", "§10", "P2", "Done", "Planned", ""),
        ("REQ-NFR-005", "NFR", "Performance PoC", "BR-008", "§9", "§10", "P3", "Done", "Planned", ""),
        ("REQ-SEC-001", "SEC", "Role labeling", "BR-006", "§8", "§11", "P1", "Done", "Planned", ""),
        ("REQ-SEC-002", "SEC", "Session termination", "BR-012", "§7", "§11", "P1", "Done", "Planned", ""),
        ("REQ-SEC-003", "SEC", "Pedagogical limitation doc", "BR-007", "§8", "§11", "P2", "Done", "Planned", ""),
    ]
    for r, row in enumerate(data, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    set_widths(ws, [12, 8, 36, 8, 12, 12, 8, 12, 12, 40])
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(data)+1}"


def sheet_tickets(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Tickets")
    headers = ["TKT_ID", "Theme", "REQ_IDs", "Sprint", "Dev_Owner", "QA_Owner", "Status", "BRD_Ref", "FRS_Ref"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    rows = [
        ("TKT-QAMVP-000", "System shell", "REQ-FR-000", "S1", "Dev", "QA", "Open", "BRD §4", "FRS §2"),
        ("TKT-QAMVP-101", "Auth happy", "REQ-FR-001,002,003", "S1", "Dev", "QA", "Open", "BRD §6", "FRS §3"),
        ("TKT-QAMVP-102", "Auth negative", "REQ-FR-004,005", "S1", "Dev", "QA", "Open", "BRD §7", "FRS §3"),
        ("TKT-QAMVP-110", "Navigation", "REQ-FR-010,011,012", "S1", "Dev", "QA", "Open", "BRD §6", "FRS §4"),
        ("TKT-QAMVP-111", "Admin nav", "REQ-FR-013", "S1", "Dev", "QA", "Open", "BRD §6", "FRS §4"),
        ("TKT-QAMVP-112", "Logout", "REQ-FR-014", "S1", "Dev", "QA", "Open", "BRD §7", "FRS §4"),
        ("TKT-QAMVP-120", "Dashboard", "REQ-FR-020,021,022,023", "S2", "Dev", "QA", "Open", "BRD §7", "FRS §5"),
        ("TKT-QAMVP-130", "Trade form A", "REQ-FR-030,031,032", "S2", "Dev", "QA", "Open", "BRD §7", "FRS §6"),
        ("TKT-QAMVP-131", "Trade submit", "REQ-FR-033,034", "S2", "Dev", "QA", "Open", "BRD §7", "FRS §6"),
        ("TKT-QAMVP-132", "Trade validation", "REQ-FR-035", "S2", "Dev", "QA", "Open", "BRD §6", "FRS §6"),
        ("TKT-QAMVP-140", "Queue view", "REQ-FR-040,041,042", "S2", "Dev", "QA", "Open", "BRD §7", "FRS §7"),
        ("TKT-QAMVP-141", "Queue approve", "REQ-FR-043,044,045", "S2", "Dev", "QA", "Open", "BRD §7", "FRS §7"),
        ("TKT-QAMVP-150", "Trade list", "REQ-FR-050,051", "S3", "Dev", "QA", "Open", "BRD §6", "FRS §8"),
        ("TKT-QAMVP-160", "Admin page", "REQ-FR-060", "S3", "Dev", "QA", "Open", "BRD §6", "FRS §9"),
        ("TKT-QAMVP-201", "NFR core", "REQ-NFR-001,002", "S3", "Dev", "QA", "Open", "BRD §4", "FRS §10"),
        ("TKT-QAMVP-202", "NFR UX", "REQ-NFR-003,004", "S3", "Dev", "QA", "Open", "BRD §4", "FRS §10"),
        ("TKT-QAMVP-203", "NFR perf", "REQ-NFR-005", "S3", "Dev", "QA", "Open", "BRD §9", "FRS §10"),
        ("TKT-QAMVP-301", "SEC session", "REQ-SEC-001,002", "S3", "Dev", "QA", "Open", "BRD §8", "FRS §11"),
        ("TKT-QAMVP-302", "SEC docs", "REQ-SEC-003", "S3", "QA", "QA", "Open", "BRD §8", "FRS §11"),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    set_widths(ws, [16, 20, 28, 8, 12, 12, 10, 12, 12])
    ws.freeze_panes = "A2"


def sheet_test_execution(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("TestExecution")
    headers = [
        "TC_ID",
        "REQ_ID",
        "TKT_ID",
        "BRD_Section",
        "FRS_Section",
        "TDS_Ref",
        "Suite",
        "Environment",
        "Last_Run",
        "Result",
        "Defect_ID",
        "Automation_Path",
    ]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    # Sample execution rows mirroring RTM
    rows = [
        ("TC-AUTH-001", "REQ-FR-001", "TKT-QAMVP-101", "BRD §6", "FRS §3.1", "TD-REQ-FR-001", "S-1", "ENV-LOCAL-01", "", "Planned", "", "test_data/generated-scripts/"),
        ("TC-AUTH-002", "REQ-FR-002", "TKT-QAMVP-101", "BRD §6", "FRS §3", "TD-REQ-FR-002", "S-1", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-AUTH-003", "REQ-FR-003", "TKT-QAMVP-101", "BRD §6", "FRS §3", "TD-REQ-FR-003", "S-1", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-AUTH-004", "REQ-FR-004", "TKT-QAMVP-102", "BRD §7", "FRS §3", "TD-REQ-FR-004", "S-3", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-AUTH-005", "REQ-FR-005", "TKT-QAMVP-102", "BRD §7", "FRS §3", "TD-REQ-FR-005", "S-3", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-NAV-003", "REQ-FR-012", "TKT-QAMVP-110", "BRD §6", "FRS §4.1", "TD-REQ-FR-012", "S-2", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-TRADE-004", "REQ-FR-033", "TKT-QAMVP-131", "BRD §7", "FRS §6", "TD-REQ-FR-033", "S-1", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-QUEUE-004", "REQ-FR-043", "TKT-QAMVP-141", "BRD §7", "FRS §7", "TD-REQ-FR-043", "S-1", "ENV-LOCAL-01", "", "Planned", "", ""),
        ("TC-DASH-001", "REQ-FR-020", "TKT-QAMVP-120", "BRD §7", "FRS §5", "TD-REQ-FR-020", "S-1", "ENV-LOCAL-01", "", "Planned", "", ""),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    set_widths(ws, [14, 12, 16, 12, 12, 18, 8, 14, 12, 10, 12, 36])
    ws.freeze_panes = "A2"


def sheet_defects(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("Defects")
    headers = ["Defect_ID", "Sev", "Summary", "REQ_ID", "TKT_ID", "Status", "Opened", "Closed"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    ws.cell(row=2, column=1, value="(none yet)")
    set_widths(ws, [12, 6, 40, 12, 14, 12, 12, 12])
    ws.freeze_panes = "A2"


def sheet_signoff(wb: openpyxl.Workbook) -> None:
    ws = wb.create_sheet("SignOff")
    headers = ["Gate", "Role", "Name", "Date", "Approved_Y_N", "Comment"]
    for i, h in enumerate(headers, 1):
        ws.cell(row=1, column=i, value=h)
    style_header(ws, len(headers))
    rows = [
        ("G-TD", "QA Lead", "", "", "", ""),
        ("G-TE", "QA Lead", "", "", "", ""),
        ("G-UAT", "Sponsor", "", "", "", ""),
    ]
    for r, row in enumerate(rows, 2):
        for c, val in enumerate(row, 1):
            ws.cell(row=r, column=c, value=val)
    set_widths(ws, [10, 14, 20, 12, 14, 40])
    ws.freeze_panes = "A2"


def main() -> None:
    wb = openpyxl.Workbook()
    sheet_instructions(wb)
    sheet_gates(wb)
    sheet_requirements(wb)
    sheet_tickets(wb)
    sheet_test_execution(wb)
    sheet_defects(wb)
    sheet_signoff(wb)
    OUT.parent.mkdir(parents=True, exist_ok=True)
    wb.save(OUT)
    print(f"Wrote {OUT}")


if __name__ == "__main__":
    main()
